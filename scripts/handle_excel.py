import os
from openpyxl import load_workbook
from scripts.handle_config import do_config
from scripts.constants import DATAS_DIR


actual_col=do_config.get_int('cases excel','actual_col')
result_col=do_config.get_int('cases excel','result_col')
mobile_sheet = do_config.get_value('mobile phone','mobile_sheet')

class DoExcel():
    """
    定义处理excel类
    """
    def __init__(self,file_name,sheet_name=None):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_cases(self):
        """
        获取所有测试用例
        :return:
        """
        wb = load_workbook(self.file_name)
        if self.sheet_name is None:
            ws = wb.active
        else:
            ws = wb[self.sheet_name]
        head_data_tuple = tuple(ws.iter_rows(max_row=1,values_only=True))[0]
        one_list = []
        for one_tuple in tuple(ws.iter_rows(min_row=2,values_only=True)):
            one_list.append(dict(zip(head_data_tuple,one_tuple)))
        return one_list

    def get_case(self,row):
        """
        获取指定某一行的用例内容
        :param row: 行号
        :return:
        """
        return self.get_cases()[row-1]
    def write_result(self,row,actual,result):
        """
        在指定的行写入数据
        :param row: 行号
        :param actual: 实际结果
        :param result: 用例执行的结果（Pass或Fail）
        :return:
        """
        wb = load_workbook(self.file_name)
        if self.sheet_name is None:
            ws = wb.active
        else:
            ws = wb[self.sheet_name]
        if isinstance(row, int) and (2 <= row <= ws.max_row):
            ws.cell(row=row, column=actual_col).value = actual
            ws.cell(row=row, column=result_col).value = result
            wb.save(self.file_name)
            wb.close()
        else:
            print('传入的行号有误！')

    def write_value(self,row,column,value,sheet_name=None):
        wb = load_workbook(self.file_name)
        if sheet_name is None:
            ws = wb.active
        else:
            ws = wb[sheet_name]
        if isinstance(row, int):
            ws.cell(row,column).value=value
            wb.save(self.file_name)
            wb.close()
        else:
            print('传入的行号或列号有误！')
    def write_register_value(self,row,select_result):
        if isinstance(select_result,dict):
            self.write_value(row,1,str(select_result['Id']),mobile_sheet)
            self.write_value(row,2,str(select_result['MobilePhone']),mobile_sheet)
            self.write_value(row,3,str(select_result['Pwd']),mobile_sheet)
            self.write_value(row,4,str(select_result['RegName']),mobile_sheet)




if __name__ == '__main__':
    # filename='./datas/test_data.xlsx'
    # do_excel=DoExcel(filename)
    # cases=do_excel.get_cases()
    # print(cases)
    # do_excel.write_result(2,'pass','Pass')
    # excel_name = do_config.get_value('file path','excel_name')
    # filename = os.path.join(DATAS_DIR,excel_name)
    # do_excel=DoExcel(filename,'subtraction')
    # cases=do_excel.get_cases()
    # print(cases)
    # do_excel.write_result(2,'pass','Pass')
    mobile_sheet = do_config.get_value('mobile phone', 'mobile_sheet')
    http_excel_path = os.path.join(DATAS_DIR, do_config.get_value('file path', 'http_excel_name'))
    mobile_phone = DoExcel(http_excel_path, mobile_sheet).get_case(1)['new_mobile_phone']
    print(mobile_phone)

